import sys 
print("Python version (should be 3.6+): {}\n". format(sys.version))

import numpy as np 
import pandas as pd 

# for text processing
import re
import nltk
from nltk.stem import WordNetLemmatizer 

# for scraping
from selenium import webdriver
from bs4 import BeautifulSoup

# for exporting to excel file
from openpyxl import load_workbook

# Excel file name
FILENAME = 'Data_Science_Internship_Assignment.xlsx'

# Ignore warnings
import warnings
warnings.filterwarnings("ignore")

##############################################
###############    Functions   ###############
##############################################

# function to clean the TAGLINE column
def process_tagline(tagline):
    """
    cleans the tagline text and converts it to tags

    Parameters:
      tagline: the text from the tagline column

    Returns: string that contains extracted tags seperated by ';'
    """

    stopwords = nltk.corpus.stopwords.words("english")
    lemmatizer = WordNetLemmatizer() 
    
    # first tokenize by sentence, then by word to ensure that punctuation is caught as it's own token
    tokens = [word for sent in nltk.sent_tokenize(tagline.lower()) for word in nltk.word_tokenize(sent)]
    
    # filter out any tokens not containing letters (e.g., numeric tokens, raw punctuation)
    filtered_tokens = []
    for token in tokens:
        if re.search('[a-zA-Z]', token):
            filtered_tokens.append(token)
    
    # lemmatize the words
    lems = [lemmatizer.lemmatize(t) for t in filtered_tokens]
    
    # remove stopwords
    final_tokens = [token for token in lems if token not in stopwords]
    
    #remove letters
    final_tokens = [token for token in final_tokens if len(token)>1]
    
    return ";".join(final_tokens)

# functions to export results to the excel file
def append_df_to_excel(filename, table, sheet_name='Sheet1', startcol=0, truncate_sheet=True):
    """
    Append a DataFrame (table) to existing Excel file into a specific Sheet.
    If Excel file or sheet doesn't exist, then this function will create it.

    Parameters:
        filename : File path or existing ExcelWriter
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame
        startcol: starting column to dump dataframe
        truncate_sheet : truncate (remove and recreate) sheet before writing dataframe to Excel file
    
    Returns: None
    """
    
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    #try:
    #    FileNotFoundError
    #except NameError:
    #    FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove sheet
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    # write out the new sheet
    table.to_excel(writer, sheet_name, index=False, startcol=startcol)

    # save the workbook
    writer.save()


##############################################
########   Entities  Classification   ########
##############################################   

print("Entities Classification...")
    
# read data
df = pd.read_excel("Data_Science_Internship_Assignment.xlsx", sheet_name = "Data")

# add launch year
df["LAUNCH YEAR"] = df["LAUNCH DATE"].str[:4]
df["LAUNCH YEAR"].fillna(df["LAUNCH DATE"], inplace=True)
df["LAUNCH YEAR"] = df["LAUNCH YEAR"].astype(int)


# combine the tags from TAGS column and processed TAGLINE column
for index, row in df.iterrows():
    df.loc[index, "ALL TAGS"] = ';'.join(set([a for a in str(row["TAGS"]).split(';') if a!= "nan"] +
                                              process_tagline(str(row["TAGLINE"])).split(';')))

# Classify the companies based on the presence of some terms in WEBSITE, NAME or ALL TAGS column
for index, row in df.iterrows():
    if (".gov" in row["WEBSITE"]) | any(s in row["ALL TAGS"] for s in ["non-profit", "charity"]):
        df.loc[index, "TYPE"] = "Government/Non-profit"
    elif (".sch" in row["WEBSITE"]) | ("School" in str(row["NAME"])):
        df.loc[index, "TYPE"] = "Universitiy/School"
    else:
        if row["LAUNCH YEAR"]<1990:
            df.loc[index, "TYPE"] = "Mature Company"
        else:
            if any(s in row["ALL TAGS"] for s in ["tech", "data", "software", "mobile", "app", "solution", "company",
                                                  "cloud", "online", "innov", "e-", "engine", "service", "web"]):
                df.loc[index, "TYPE"] = "Startup"
            else:
                df.loc[index, "TYPE"] = "Unclassified"

# export to excel file      
append_df_to_excel(FILENAME, df[["TYPE"]], "Data", startcol= 10, truncate_sheet = False)
append_df_to_excel(FILENAME, pd.DataFrame(df.TYPE.value_counts()).reset_index().rename(
    {"TYPE":"Count", "index":"Entity"}, axis='columns'), "Count")
                
print("\nSuccessefully exported entities to Excel file.")

##############################################
#############   Web  Scrapping   #############
##############################################                                             
                        
print("\nNote: a website will open automatically in a new chrome window. Don't panic, it's just part of the code :)")
print("\nWeb Scrapping...")

# load the website content
driver = webdriver.Chrome("/Users/aminemajdoubi/Downloads/chromedriver") # download chromedriver and insert right path

# since the website needs time to load, you need to make the driver implicitly wait for 5 seconds and then reopen link
driver.get("https://www.ycombinator.com/companies/")
driver.implicitly_wait(5)
driver.get("https://www.ycombinator.com/companies/")

# load web page content to BeautifulSoup                                       
soup = BeautifulSoup(driver.page_source)

names = []
years = []
descriptions = []
links = []
seasons = []

# create a table of names, links, descriptions and Launch years/seasons of the companies in the website                                         
for a in soup.find("table").find_all("tr"):
    content = a.find_all("td")
    names.append(content[0].text)
    years.append(content[1].text[1:])
    seasons.append(content[1].text[0])
    descriptions.append(content[2].text)
    try:
        links.append(a.find("a",href=True)["href"])
    except:
        links.append("")
        
# export to excel file
df_scraping = pd.DataFrame({'Name':names, 'Year':years, 'Season':seasons, 'Description':descriptions, 'Link':links})
append_df_to_excel(FILENAME, df_scraping, "Scraping results")

print("\nSuccessefully exported web content to Excel file.\n")
